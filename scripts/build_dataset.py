#!/usr/bin/env python3
"""Build normalized thermo lookup tables from an .xlsm workbook."""

from __future__ import annotations

import argparse
import datetime as dt
import json
import math
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import load_workbook

SATURATION_PROPS = [
    "vf",
    "vfg",
    "vg",
    "uf",
    "ufg",
    "ug",
    "hf",
    "hfg",
    "hg",
    "sf",
    "sfg",
    "sg",
]
GRID_PROPS = ["v", "u", "h", "s"]
ALL_PROP_KEYS = set(SATURATION_PROPS + GRID_PROPS + ["T", "P"])


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    return re.sub(r"\s+", " ", text)


def normalize_text(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()


def to_number(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        num = float(value)
        if math.isfinite(num):
            return num
        return None
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", "")
    try:
        num = float(text)
        if math.isfinite(num):
            return num
    except ValueError:
        return None
    return None


def infer_unit_system(sheet_name: str) -> str:
    return "ENG" if sheet_name.upper().startswith("ENG_") else "SI"


def infer_fluid(sheet_name: str) -> str:
    text = sheet_name
    text = re.sub(r"(?i)^eng_", "", text)
    text = re.sub(r"(?i)^s\.w\.\s*", "", text)
    text = re.sub(r"(?i)\b(superheated|saturated|compressed|solid|vapor|liquid|entry)\b", "", text)
    text = re.sub(r"[_-]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text or sheet_name


def choose_header_row(rows: List[List[Any]]) -> int:
    """Pick the most likely header row from the top of the sheet."""
    keywords = [
        "temp",
        "temperature",
        "pressure",
        "volume",
        "internal",
        "energy",
        "enthalpy",
        "entropy",
        "sat",
    ]
    best_idx = 0
    best_score = float("-inf")

    for idx, row in enumerate(rows[:12]):
        texts = [clean_text(cell) for cell in row if clean_text(cell)]
        if not texts:
            continue
        joined = " ".join(texts).lower()
        keyword_hits = sum(1 for key in keywords if key in joined)
        numeric_cells = sum(1 for cell in row if to_number(cell) is not None)
        score = keyword_hits * 4 + len(texts) - numeric_cells * 2
        if score > best_score:
            best_idx = idx
            best_score = score

    return best_idx


def header_text_for_column(rows: List[List[Any]], header_idx: int, col_idx: int) -> str:
    pieces: List[str] = []
    start = max(0, header_idx - 1)
    end = min(len(rows), header_idx + 3)
    for ridx in range(start, end):
        row = rows[ridx]
        value = row[col_idx] if col_idx < len(row) else None
        text = clean_text(value)
        if text and text not in pieces:
            pieces.append(text)
    return " | ".join(pieces)


def map_property(header_text: str) -> Optional[str]:
    raw = header_text.lower()
    norm = normalize_text(raw)

    token_hits = re.findall(r"\(\s*([a-z0-9]+)\s*\)", raw)
    for token in token_hits:
        if token in ALL_PROP_KEYS:
            return token

    # Saturated-table property columns (more specific first).
    if re.search(r"\bvfg\b|\(\s*vfg\s*\)|evap(oration)?\s*\(\s*vfg\s*\)", raw):
        return "vfg"
    if re.search(r"\bufg\b|\(\s*ufg\s*\)|evap(oration)?\s*\(\s*ufg\s*\)", raw):
        return "ufg"
    if re.search(r"\bhfg\b|\(\s*hfg\s*\)|evap(oration)?\s*\(\s*hfg\s*\)", raw):
        return "hfg"
    if re.search(r"\bsfg\b|\(\s*sfg\s*\)|evap(oration)?\s*\(\s*sfg\s*\)", raw):
        return "sfg"

    if re.search(r"\bvf\b|\(\s*vf\s*\)|sat(\.|urated)?\s*(liquid|solid).*\(\s*vf\s*\)", raw):
        return "vf"
    if re.search(r"\bvg\b|\(\s*vg\s*\)", raw):
        return "vg"
    if re.search(r"\buf\b|\(\s*uf\s*\)|sat(\.|urated)?\s*(liquid|solid).*\(\s*uf\s*\)", raw):
        return "uf"
    if re.search(r"\bug\b|\(\s*ug\s*\)|sat(\.|urated)?\s*vap(or)?\s*\(\s*ug\s*\)", raw) and "ufg" not in raw:
        return "ug"
    if re.search(r"\bhf\b|\(\s*hf\s*\)|sat(\.|urated)?\s*(liquid|solid).*\(\s*hf\s*\)", raw):
        return "hf"
    if re.search(r"\bhg\b|\(\s*hg\s*\)|sat(\.|urated)?\s*vap(or)?\s*\(\s*hg\s*\)", raw) and "hfg" not in raw:
        return "hg"
    if re.search(r"\bsf\b|\(\s*sf\s*\)|sat(\.|urated)?\s*(liquid|solid).*\(\s*sf\s*\)", raw):
        return "sf"
    if re.search(r"\bsg\b|\(\s*sg\s*\)|sat(\.|urated)?\s*vap(or)?\s*\(\s*sg\s*\)", raw) and "sfg" not in raw:
        return "sg"

    # Index columns.
    if re.search(r"\b(temp|temperature|tsat)\b", norm):
        return "T"
    if re.search(r"\b(press|pressure|psat)\b", norm):
        return "P"

    # Generic single-phase columns.
    if re.search(r"\bvolume\b", norm) and re.search(r"\bv\b", norm):
        return "v"
    if re.search(r"\binternal\s+energy\b", norm) and re.search(r"\bu\b", norm):
        return "u"
    if re.search(r"\benthalpy\b", norm) and re.search(r"\bh\b", norm):
        return "h"
    if re.search(r"\bentropy\b", norm) and re.search(r"\bs\b", norm):
        return "s"

    # Fallback short names.
    if norm == "v":
        return "v"
    if norm == "u":
        return "u"
    if norm == "h":
        return "h"
    if norm == "s":
        return "s"
    if norm == "temperature":
        return "T"
    if norm == "pressure":
        return "P"

    return None


def map_columns(rows: List[List[Any]], header_idx: int) -> Dict[str, int]:
    top_rows = rows[: min(len(rows), header_idx + 3)]
    if not top_rows:
        return {}
    max_cols = max(len(r) for r in top_rows)

    mapped: Dict[str, int] = {}
    for col_idx in range(max_cols):
        header = header_text_for_column(rows, header_idx, col_idx)
        prop = map_property(header)
        if prop and prop not in mapped:
            mapped[prop] = col_idx
    return mapped


def detect_mode(sheet_name: str, mapped: Dict[str, int]) -> Optional[str]:
    if "T" not in mapped or "P" not in mapped:
        return None

    sat_count = sum(1 for key in SATURATION_PROPS if key in mapped)
    grid_count = sum(1 for key in GRID_PROPS if key in mapped)
    if sat_count >= 3:
        name_low = sheet_name.lower()
        p_first = mapped["P"] < mapped["T"]
        if "pressure" in name_low or p_first:
            return "sat-P"
        return "sat-T"
    if grid_count >= 1:
        return "PT"
    return None


def iter_sheet_rows(ws: Any) -> List[List[Any]]:
    rows: List[List[Any]] = []
    for row in ws.iter_rows(values_only=True):
        values = list(row)
        while values and (values[-1] is None or str(values[-1]).strip() == ""):
            values.pop()
        rows.append(values)
    return rows


def extract_records(rows: List[List[Any]], mapped: Dict[str, int], mode: str, header_idx: int) -> List[Dict[str, float]]:
    records: List[Dict[str, float]] = []
    last_index_values: Dict[str, float] = {}

    for row in rows[header_idx + 1 :]:
        parsed: Dict[str, float] = {}
        for key, col_idx in mapped.items():
            if col_idx >= len(row):
                continue
            number = to_number(row[col_idx])
            if number is not None:
                parsed[key] = number

        index_keys: List[str]
        if mode == "PT":
            index_keys = ["P"]
        elif mode == "sat-T":
            index_keys = ["T"]
        else:
            index_keys = ["P"]

        for idx_key in index_keys:
            if idx_key in parsed:
                last_index_values[idx_key] = parsed[idx_key]
            elif idx_key in last_index_values:
                parsed[idx_key] = last_index_values[idx_key]

        if mode == "PT":
            if "T" not in parsed or "P" not in parsed:
                continue
            if not any(prop in parsed for prop in GRID_PROPS):
                continue
        elif mode == "sat-T":
            if "T" not in parsed:
                continue
            if not any(prop in parsed for prop in SATURATION_PROPS + ["P"]):
                continue
        elif mode == "sat-P":
            if "P" not in parsed:
                continue
            if not any(prop in parsed for prop in SATURATION_PROPS + ["T"]):
                continue

        records.append(parsed)

    return records


def dedupe_and_sort(records: List[Dict[str, float]], mode: str) -> List[Dict[str, float]]:
    deduped: List[Dict[str, float]] = []
    seen = set()

    for row in records:
        if mode == "PT":
            key = (round(row.get("P", 0.0), 9), round(row.get("T", 0.0), 9))
        elif mode == "sat-T":
            key = ("T", round(row.get("T", 0.0), 9))
        else:
            key = ("P", round(row.get("P", 0.0), 9))
        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)

    if mode == "PT":
        deduped.sort(key=lambda r: (r.get("P", float("inf")), r.get("T", float("inf"))))
    elif mode == "sat-T":
        deduped.sort(key=lambda r: r.get("T", float("inf")))
    else:
        deduped.sort(key=lambda r: r.get("P", float("inf")))
    return deduped


def make_table_id(sheet_name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", sheet_name.lower()).strip("-")


def build_dataset(input_path: Path) -> Dict[str, Any]:
    wb = load_workbook(input_path, data_only=True, read_only=True)
    tables: List[Dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = iter_sheet_rows(ws)
        if not rows:
            continue

        header_idx = choose_header_row(rows)
        mapped = map_columns(rows, header_idx)
        mode = detect_mode(sheet_name, mapped)
        if not mode:
            continue

        records = extract_records(rows, mapped, mode, header_idx)
        records = dedupe_and_sort(records, mode)
        if len(records) < 3:
            continue

        available_props = sorted({k for row in records for k in row.keys() if k not in {"T", "P"}})
        if not available_props:
            continue

        table: Dict[str, Any] = {
            "id": make_table_id(sheet_name),
            "sheet_name": sheet_name,
            "fluid": infer_fluid(sheet_name),
            "unit_system": infer_unit_system(sheet_name),
            "mode": mode,
            "columns": mapped,
            "properties": available_props,
            "row_count": len(records),
            "rows": records,
        }

        if mode == "PT":
            t_values = [r["T"] for r in records if "T" in r]
            p_values = [r["P"] for r in records if "P" in r]
            table["inputs"] = {
                "T": {"min": min(t_values), "max": max(t_values)},
                "P": {"min": min(p_values), "max": max(p_values)},
            }
        elif mode == "sat-T":
            values = [r["T"] for r in records if "T" in r]
            table["inputs"] = {"T": {"min": min(values), "max": max(values)}}
        else:
            values = [r["P"] for r in records if "P" in r]
            table["inputs"] = {"P": {"min": min(values), "max": max(values)}}

        tables.append(table)

    tables.sort(key=lambda t: (t["fluid"].lower(), t["sheet_name"].lower()))
    return {
        "generated_at_utc": dt.datetime.now(dt.UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "source_workbook": input_path.name,
        "table_count": len(tables),
        "tables": tables,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Normalize thermo tables from workbook to JSON.")
    parser.add_argument(
        "--input",
        default="Themodynamic and Transport Properties.xlsm",
        help="Path to source .xlsm workbook.",
    )
    parser.add_argument(
        "--output",
        default="data/thermo_tables.json",
        help="Path to write normalized JSON dataset.",
    )
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    dataset = build_dataset(input_path)
    output_path.write_text(json.dumps(dataset, indent=2), encoding="utf-8")
    print(f"Wrote {dataset['table_count']} tables to {output_path}")


if __name__ == "__main__":
    main()
